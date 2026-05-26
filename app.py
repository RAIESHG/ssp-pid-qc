"""
SSP Engineering & Consulting — P&ID QC Review Tool
Streamlit Web App  |  Powered by Google Gemini 2.5 Flash
"""

import streamlit as st
import os, io, json, time, tempfile
from datetime import date
from pathlib import Path
import pandas as pd

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SSP P&ID QC Review",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── STYLING ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background:#0D0F11; }
[data-testid="stSidebar"]          { background:#141618; }
[data-testid="stHeader"]           { background:transparent; }

.ssp-header {
    background: linear-gradient(90deg, #141618 0%, #1C1F23 100%);
    border-left: 5px solid #E8A020;
    border-radius: 6px;
    padding: 18px 24px;
    margin-bottom: 24px;
}
.ssp-header h1 { color:#E8A020; margin:0; font-size:1.6rem; }
.ssp-header p  { color:#8D9399; margin:4px 0 0; font-size:.9rem; }

.sev-card {
    border-radius: 6px; padding: 14px 10px;
    text-align: center; font-weight: bold;
}
.sev-HIGH   { background:#3D0A0A; border:1px solid #E84040; color:#E84040; }
.sev-MEDIUM { background:#2E2000; border:1px solid #E8B020; color:#E8B020; }
.sev-LOW    { background:#0A1A2E; border:1px solid #3A9FFF; color:#3A9FFF; }
.sev-INFO   { background:#0A1F14; border:1px solid #2DCA72; color:#2DCA72; }
.sev-TOTAL  { background:#1C1F23; border:1px solid #E8A020; color:#E8A020; }
.sev-num    { font-size:2.4rem; line-height:1.1; }
.sev-lbl    { font-size:.75rem; letter-spacing:.08em; margin-top:4px; }

div[data-testid="stProgress"] > div > div { background:#E8A020 !important; }
</style>
""", unsafe_allow_html=True)

# ─── DEPENDENCIES ─────────────────────────────────────────────────────────────
@st.cache_resource
def load_deps():
    missing = []
    for pkg, imp in [("google-genai","google.genai"),
                     ("pdf2image","pdf2image"),
                     ("Pillow","PIL"),
                     ("openpyxl","openpyxl")]:
        try: __import__(imp)
        except ImportError: missing.append(pkg)
    return missing

missing = load_deps()
if missing:
    st.error(f"Missing packages: `pip install {' '.join(missing)}`")
    st.stop()

from google import genai
from google.genai import types
from pdf2image import convert_from_path
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
# API key — secrets/env only, never exposed to UI
def _get_secret_key():
    try:
        return st.secrets["GEMINI_API_KEY"]
    except Exception:
        return os.environ.get("GEMINI_API_KEY", "")

# Model fallback chain — tried in order when quota/rate limit is hit
MODELS = [
    "gemini-2.5-flash",
    "gemini-3-flash-preview",
    "gemini-2.0-flash",
]
MODEL = MODELS[0]   # tracks active model during a run

MAX_TILE_DIM    = 2200
TILE_OVERLAP    = 0.10
API_DELAY       = 7.0

SEV_ORDER  = ["HIGH", "MEDIUM", "LOW", "INFO"]
SEV_COLORS = {
    "HIGH":   "#E84040",
    "MEDIUM": "#E8B020",
    "LOW":    "#3A9FFF",
    "INFO":   "#2DCA72",
}

# ─── CORE: PDF → TILES ────────────────────────────────────────────────────────
def pdf_to_tiles(pdf_bytes: bytes, dpi: int = 250):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp = f.name
    try:
        pages = convert_from_path(tmp, dpi=dpi)
    finally:
        os.unlink(tmp)

    tiles = []
    for pn, page in enumerate(pages, 1):
        W, H = page.size
        ratio    = min(MAX_TILE_DIM/W, MAX_TILE_DIM/H)
        overview = page.resize((int(W*ratio), int(H*ratio)), Image.LANCZOS)
        tiles.append((pn, "full-page overview", overview))

        ox, oy = int(W*TILE_OVERLAP), int(H*TILE_OVERLAP)
        for label, x1, y1, x2, y2 in [
            ("top-left",     0,         0,         W//2+ox, H//2+oy),
            ("top-right",    W//2-ox,   0,         W,       H//2+oy),
            ("bottom-left",  0,         H//2-oy,   W//2+ox, H),
            ("bottom-right", W//2-ox,   H//2-oy,   W,       H),
        ]:
            tile = page.crop((max(0,x1), max(0,y1), min(W,x2), min(H,y2)))
            tw, th = tile.size
            if max(tw, th) > MAX_TILE_DIM:
                r = MAX_TILE_DIM / max(tw, th)
                tile = tile.resize((int(tw*r), int(th*r)), Image.LANCZOS)
            tiles.append((pn, label, tile))

    return tiles, len(pages)

def pdf_to_full_pages(pdf_bytes: bytes, dpi: int = 250):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp = f.name
    try:
        pages = convert_from_path(tmp, dpi=dpi)
    finally:
        os.unlink(tmp)
    return [(pn, page) for pn, page in enumerate(pages, 1)]

def pil_to_bytes(img: Image.Image) -> bytes:
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=93)
    return buf.getvalue()

# ─── CORE: PROMPT ─────────────────────────────────────────────────────────────
def build_prompt(page_num, total_pages, filename):
    return f"""You are a senior P&ID engineer at SSP Engineering & Consulting performing a QC review.

DRAWING: {filename} | PAGE: {page_num}/{total_pages}

READ EVERY PIECE OF TEXT including very small annotations. Name exact element locations.

━━━ EXTRACT ALL OF THE FOLLOWING ━━━
1. EQUIPMENT TAGS — TK-xxxx, P-xxxx, V-xxxx, E-xxxx, C-xxxx
2. LINE NUMBERS — every line number (format SIZE\"-FLUID-SEQ-SPEC-TRACE)
3. INSTRUMENT TAGS — FIC, LIT, AIT, PSV, XV, PCV, HS, ZSC, ZSO, ZLO, ZLC, PAH, PSH, KDY, SY, SC, FIT, FI, LI, AE, AI, PG, XY, XL
4. VALVE TAGS — XV, HV, MOV, PCV, PSV, PVRV, FCV
5. PIPE SPECS — e.g. A520-N, A130-N, A021-N
6. PLACEHOLDER SIZES — every X\" or X\"xX\" instance with its exact location

━━━ QC CHECKS ━━━

TAGGING:
• Instrument bubble with NO tag number → MEDIUM
• XV or MOV with NO individual number → HIGH
• PSV/PRV/PVRV with NO tag number → HIGH
• PSV with no set pressure → HIGH
• PCV/FCV with no loop number → MEDIUM
• PAH/PSH without individual tag numbers → MEDIUM

PIPING:
• X\" size on XS/acid service line → HIGH
• X\" size on utility line → MEDIUM
• X\"xX\" reducer callout → MEDIUM
• Line missing pipe spec → MEDIUM

DATASHEETS:
• DESIGN PRESSURE or DESIGN TEMP blank → HIGH
• SEAL PLAN blank on acid service pump → HIGH
• SIZE, VOLUME, MOC, RATED CAPACITY, RATED POWER, DP, INSULATION blank → MEDIUM

SAFETY:
• PVRV on acid tank — no tag or no set pressure → HIGH
• PSV on blocked discharge — no tag number → HIGH

DRAWING/NOTES:
• DESIGN BY blank, SCALE \"NONE\" not \"NTS\", PLOT DATE \"----\" → LOW
• Note with \"00XX\" placeholder → LOW
• Control valve sizes PRELIMINARY → LOW
• System shown as TYPICAL → INFO

━━━ SEVERITY ━━━
HIGH   = Untagged PSV/PVRV/XV on acid | Blank design pressure/temp | X\" on acid line | Blank seal plan
MEDIUM = Untagged instrument | PCV no loop | X\" utility | Incomplete datasheet
LOW    = Title block / note omission
INFO   = Stage observation

━━━ CRITICAL: PIXEL COORDINATES ━━━
For EACH issue found, provide the pixel coordinates (x, y) where the error is located.
x ranges from 0 (left edge) to image_width (right edge)
y ranges from 0 (top edge) to image_height (bottom edge)
Place coordinates at the CENTER of the problematic element.

Return ONLY raw valid JSON, no markdown, no fences:
{{"extracted":{{"equipment_tags":[],"line_numbers":[],"instrument_tags":[],"valve_tags":[],"pipe_specs":[],"placeholder_sizes":[]}},"issues":[{{"severity":"HIGH","category":"Tagging","element":"exact element","issue":"specific problem","recommendation":"specific fix","x":500,"y":300}}]}}

If page is blank/border only: {{"extracted":{{"equipment_tags":[],"line_numbers":[],"instrument_tags":[],"valve_tags":[],"pipe_specs":[],"placeholder_sizes":[]}},"issues":[]}}"""

# ─── CORE: GEMINI API CALL ────────────────────────────────────────────────────
def call_gemini(client, img_bytes, prompt, retries=2):
    """Try each model in MODELS; on quota exhaustion fall through to the next."""
    global MODEL
    empty = {"extracted":{}, "issues":[]}

    for model in MODELS:
        MODEL = model   # keep track of active model for display/export
        for attempt in range(retries):
            try:
                resp = client.models.generate_content(
                    model=model,
                    contents=[
                        types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"),
                        types.Part.from_text(text=prompt),
                    ]
                )
                raw = resp.text.strip().replace("```json","").replace("```","").strip()
                s = raw.find("{"); e = raw.rfind("}")
                if s == -1: raise ValueError("No JSON in response")
                return json.loads(raw[s:e+1])

            except json.JSONDecodeError:
                if attempt < retries - 1:
                    time.sleep(4)
                else:
                    return empty   # bad JSON even after retry — skip tile

            except Exception as ex:
                msg = str(ex)
                is_quota = ("429" in msg or "quota" in msg.lower()
                            or "exhausted" in msg.lower() or "RESOURCE_EXHAUSTED" in msg)
                if is_quota:
                    # Break inner loop → try next model
                    break
                elif attempt < retries - 1:
                    time.sleep(5)
                else:
                    return empty   # non-quota error, give up

    return empty   # all models exhausted

# ─── CORE: DEDUPLICATION ──────────────────────────────────────────────────────
def dedupe_issues(issues):
    seen, out = set(), []
    for iss in issues:
        key = (str(iss.get("element","")) + str(iss.get("issue",""))[:40]).lower().strip()
        if key not in seen and iss.get("issue"):
            seen.add(key); out.append(iss)
    return out

def dedupe_list(lst):
    return sorted({str(x).strip() for x in lst if str(x).strip()})

# ─── CORE: EXCEL EXPORT (returns bytes) ───────────────────────────────────────
def export_excel_bytes(issues, extracted, meta):
    wb  = openpyxl.Workbook()
    SEV = {"HIGH":("3D0A0A","E84040"),"MEDIUM":("2E2000","E8B020"),
           "LOW":("0A1A2E","3A9FFF"),"INFO":("0A1F14","2DCA72")}
    sev_counts = {s: sum(1 for i in issues if i.get("severity")==s) for s in SEV}

    def fill(c):                return PatternFill("solid", fgColor=c)
    def font(c,bold=False,sz=10): return Font(name="Courier New",color=c,bold=bold,size=sz)
    def bdr():
        s = Side(style="thin", color="2A2E33")
        return Border(left=s, right=s, top=s, bottom=s)
    def align(wrap=False):      return Alignment(vertical="top", wrap_text=wrap)

    # Sheet 1: QC Issues
    ws = wb.active; ws.title = "QC Issues"; ws.sheet_properties.tabColor = "E84040"
    ws.merge_cells("A1:G1")
    ws["A1"] = "SSP Engineering & Consulting — P&ID QC Review Report"
    ws["A1"].font = font("E8A020", bold=True, sz=14)
    ws["A1"].fill = fill("141618")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    meta_items = [("Drawing",meta.get("drawing","—")),("Revision",meta.get("revision","—")),
                  ("Project",meta.get("project","—")),("Stage",meta.get("stage","—")),
                  ("Client",meta.get("client","—")),("Reviewer",meta.get("reviewer","—")),
                  ("Date",str(date.today())),("Model",MODEL)]
    for idx,(k,v) in enumerate(meta_items):
        row=2+idx//4; col=1+(idx%4)*2
        ck=ws.cell(row=row,column=col,value=k)
        cv=ws.cell(row=row,column=col+1,value=v)
        ck.font=font("8D9399"); ck.fill=fill("1C1F23"); ck.border=bdr()
        cv.font=font("F5C060"); cv.fill=fill("141618"); cv.border=bdr()

    for col,(sev,cnt) in enumerate(sev_counts.items(),1):
        bg,fg=SEV[sev]; c=ws.cell(row=4,column=col,value=f"{sev}: {cnt}")
        c.font=font(fg,bold=True); c.fill=fill(bg); c.border=bdr()
    c=ws.cell(row=4,column=5,value=f"TOTAL: {len(issues)}")
    c.font=font("E2E5E8",bold=True); c.fill=fill("141618"); c.border=bdr()
    ws.row_dimensions[4].height=18

    headers=["#","SEVERITY","CATEGORY","ELEMENT","ISSUE","RECOMMENDATION","PAGE"]
    widths =[5,   12,        16,         30,       62,     62,              8]
    HR=5
    for col,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=HR,column=col,value=h)
        c.font=font("8D9399",bold=True); c.fill=fill("1C1F23"); c.border=bdr()
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes=f"A{HR+1}"
    ws.auto_filter.ref=f"A{HR}:G{HR+len(issues)}"

    for ri,iss in enumerate(issues):
        row=HR+1+ri; sev=iss.get("severity","INFO")
        bg,fg=SEV.get(sev,("141618","E2E5E8"))
        ws.row_dimensions[row].height=55
        vals=[ri+1,sev,iss.get("category",""),iss.get("element",""),
              iss.get("issue",""),iss.get("recommendation",""),iss.get("page",1)]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=row,column=col,value=val); c.border=bdr(); c.alignment=align(wrap=True)
            if col==2:   c.font=font(fg,bold=True); c.fill=fill(bg)
            elif col==4: c.font=font("F5C060");     c.fill=fill("141618")
            elif col==6: c.font=font("8D9399");     c.fill=fill("141618")
            else:        c.font=font("E2E5E8");     c.fill=fill("141618")

    # Sheet 2: Extracted Data
    ws2=wb.create_sheet("Extracted Data"); ws2.sheet_properties.tabColor="2DCA72"
    ws2.merge_cells("A1:B1")
    ws2["A1"]=f"Extracted Data — {meta.get('drawing','')} {meta.get('revision','')}"
    ws2["A1"].font=font("E8A020",bold=True,sz=13); ws2["A1"].fill=fill("141618")
    ws2.column_dimensions["A"].width=28; ws2.column_dimensions["B"].width=65
    row=2
    for title,key,fg_c,bg_c in [
        ("Equipment Tags","equipment_tags","2DCA72","0A1F14"),
        ("Line Numbers","line_numbers","2DCA72","0A1F14"),
        ("Instrument Tags","instrument_tags","2DCA72","0A1F14"),
        ("Valve Tags","valve_tags","2DCA72","0A1F14"),
        ("Pipe Specs","pipe_specs","2DCA72","0A1F14"),
        ('Placeholder X" Sizes',"placeholder_sizes","E8B020","2E2000"),
    ]:
        items=extracted.get(key,[])
        for c in [ws2.cell(row=row,column=1,value=title),
                  ws2.cell(row=row,column=2,value=f"{len(items)} found")]:
            c.font=font("8D9399",bold=True); c.fill=fill("1C1F23"); c.border=bdr()
        row+=1
        for item in items:
            ws2.cell(row=row,column=1).fill=fill("141618")
            v=ws2.cell(row=row,column=2,value=str(item))
            v.font=font(fg_c); v.fill=fill(bg_c); v.border=bdr(); row+=1

    # Sheet 3: Summary
    ws3=wb.create_sheet("Summary"); ws3.sheet_properties.tabColor="E8A020"
    ws3.column_dimensions["A"].width=24; ws3.column_dimensions["B"].width=48
    sev_fg={"HIGH":"E84040","MEDIUM":"E8B020","LOW":"3A9FFF","INFO":"2DCA72","TOTAL":"E8A020"}
    for ri,(k,v) in enumerate([
        ("SSP P&ID QC REPORT",""),("",""),
        ("Drawing",meta.get("drawing","—")),("Revision",meta.get("revision","—")),
        ("Project",meta.get("project","—")),("Stage",meta.get("stage","—")),
        ("Client",meta.get("client","—")),("Reviewer",meta.get("reviewer","—")),
        ("Date",str(date.today())),("Model",MODEL),
        ("",""),("ISSUE COUNTS",""),
        ("HIGH",sev_counts["HIGH"]),("MEDIUM",sev_counts["MEDIUM"]),
        ("LOW",sev_counts["LOW"]),("INFO",sev_counts["INFO"]),
        ("TOTAL",len(issues)),("",""),("EXTRACTED",""),
        ("Equipment Tags",len(extracted.get("equipment_tags",[]))),
        ("Line Numbers",len(extracted.get("line_numbers",[]))),
        ("Instrument Tags",len(extracted.get("instrument_tags",[]))),
        ("Placeholder Sizes",len(extracted.get("placeholder_sizes",[]))),
    ],1):
        is_hdr=k in("SSP P&ID QC REPORT","ISSUE COUNTS","EXTRACTED")
        ck=ws3.cell(row=ri,column=1,value=k); cv=ws3.cell(row=ri,column=2,value=v)
        fg=sev_fg.get(k,"8D9399" if not k else "E2E5E8")
        ck.font=font("E8A020" if is_hdr else fg,bold=is_hdr,sz=13 if is_hdr else 10)
        cv.font=font(fg,bold=k in sev_fg)
        for c in[ck,cv]:
            c.fill=fill("1C1F23" if is_hdr else "141618")
            if k: c.border=bdr()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─── OUTPUT: MARKED-UP PDF EXPORT (returns bytes) ─────────────────────────────
def export_pdf_markup_bytes(issues, original_pdf_bytes, dpi=250):
    """Draw circles at exact error coordinates on PDF pages."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(original_pdf_bytes)
        tmp = f.name
    try:
        pages = convert_from_path(tmp, dpi=dpi)
    finally:
        try: os.unlink(tmp)
        except Exception: pass

    out_images = []
    for pn, img in enumerate(pages, start=1):
        draw = ImageDraw.Draw(img, "RGBA")
        page_issues = [i for i in issues if int(i.get("page", 1)) == pn]

        for iss in page_issues:
            x = iss.get("x")
            y = iss.get("y")
            if x is not None and y is not None:
                sev = (iss.get("severity") or "INFO").upper()
                color = SEV_COLORS.get(sev, "#2DCA72")
                x, y = int(x), int(y)
                radius = 40

                draw.ellipse([(x-radius, y-radius), (x+radius, y+radius)],
                            outline=color, width=4)
                draw.ellipse([(x-radius+3, y-radius+3), (x+radius-3, y+radius-3)],
                            outline=color, width=1)

        out_images.append(img.convert("RGB"))

    buf = io.BytesIO()
    if out_images:
        out_images[0].save(buf, format="PDF", save_all=True, append_images=out_images[1:])
    else:
        return original_pdf_bytes
    buf.seek(0)
    return buf.getvalue()

# ─── SESSION STATE INIT ───────────────────────────────────────────────────────
for key, default in [
    ("results", None),
    ("log",     []),
    ("running", False),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="ssp-header">
  <h1>⚙️ SSP Engineering &amp; Consulting</h1>
  <p>P&amp;ID QC Review Tool &nbsp;·&nbsp; Powered by Google Gemini 2.5 Flash</p>
</div>
""", unsafe_allow_html=True)

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")

    _secret_key = _get_secret_key()
    if _secret_key:
        api_key = _secret_key
        st.caption("🔒 API key configured.")
    else:
        api_key = st.text_input(
            "Gemini API Key",
            type="password",
            placeholder="Paste your Gemini API key",
            help="Get a free key at aistudio.google.com/app/apikey",
        )

    st.markdown("---")
    st.markdown("### 📋 Drawing Metadata")
    drawing  = st.text_input("Drawing Number",  placeholder="e.g. 1010-PID-0011")
    revision = st.text_input("Revision",         placeholder="e.g. B")
    project  = st.text_input("Project Number",   placeholder="e.g. 25MA02")
    client_n = st.text_input("Client",           placeholder="e.g. Motiva Enterprises")
    reviewer = st.text_input("Reviewed By",      placeholder="Your name")
    stage    = st.text_input("Design Stage",     placeholder="e.g. FEL2B", value="FEL2B")

    st.markdown("---")
    st.markdown("### 🔧 Advanced")
    dpi = st.slider("Render DPI", min_value=150, max_value=400, value=250, step=50,
                    help="Higher = more detail but slower. 250 is optimal for P&IDs.")

# ─── MAIN: FILE UPLOAD ────────────────────────────────────────────────────────
col_upload, col_info = st.columns([2, 1])

with col_upload:
    uploaded = st.file_uploader(
        "Upload P&ID Drawing (PDF)",
        type=["pdf"],
        help="Multi-page PDFs supported — each page is analyzed independently.",
    )

with col_info:
    if uploaded:
        st.success(f"**{uploaded.name}**  \n{uploaded.size/1024:.0f} KB")
    else:
        st.info("Upload a PDF to begin the QC review.")

# ─── MAIN: RUN BUTTON ─────────────────────────────────────────────────────────
st.markdown("---")

run_col, _ = st.columns([1, 3])
with run_col:
    run_btn = st.button(
        "▶ Run QC Analysis",
        type="primary",
        disabled=(uploaded is None or st.session_state.running),
        use_container_width=True,
    )

# ─── MAIN: ANALYSIS ENGINE ────────────────────────────────────────────────────
if run_btn and uploaded:
    st.session_state.running = True
    st.session_state.results = None
    st.session_state.log     = []

    meta = dict(
        drawing  = drawing  or Path(uploaded.name).stem,
        revision = revision or "—",
        project  = project  or "—",
        client   = client_n or "—",
        reviewer = reviewer or "—",
        stage    = stage    or "FEL2B",
    )

    status_box   = st.empty()
    progress_bar = st.progress(0)
    log_box      = st.empty()

    def log(msg):
        st.session_state.log.append(msg)
        log_box.markdown(
            "<br>".join(st.session_state.log[-12:]),
            unsafe_allow_html=True,
        )

    try:
        # 1. Connect to Gemini — test primary model, fall back silently if needed
        status_box.info("🔌 Connecting to Gemini…")
        log("🔌 Connecting to Gemini API…")
        client = genai.Client(api_key=api_key)
        active_model = None
        for m in MODELS:
            try:
                client.models.generate_content(model=m, contents="Reply OK only.")
                active_model = m
                break
            except Exception as ex:
                if "429" in str(ex) or "quota" in str(ex).lower() or "exhausted" in str(ex).lower():
                    log(f"⚠ `{m}` quota exhausted — trying next model…")
                else:
                    raise
        if not active_model:
            raise RuntimeError("All models exhausted or unavailable. Try again later.")
        global MODEL
        MODEL = active_model
        log(f"✅ Connected — model: `{MODEL}`")

        # 2. Render PDF
        status_box.info("📄 Rendering PDF…")
        log(f"📄 Rendering PDF at {dpi} DPI…")
        pdf_bytes = uploaded.read()
        full_pages = pdf_to_full_pages(pdf_bytes, dpi=dpi)
        total_pages = len(full_pages)
        log(f"✅ {total_pages} page(s) to analyze")
        est = total_pages * API_DELAY
        log(f"⏱ Estimated time: ~{est/60:.1f} min (free-tier pacing)")

        # 3. Analyze full pages
        status_box.info("🔍 Analyzing pages…")
        all_issues = []
        all_ext    = {k:[] for k in ["equipment_tags","line_numbers","instrument_tags",
                                      "valve_tags","pipe_specs","placeholder_sizes"]}

        for idx, (page_num, page_img) in enumerate(full_pages):
            pct = idx / total_pages
            progress_bar.progress(pct)

            log(f"<br>🖼 [{idx+1}/{total_pages}] Page {page_num} {page_img.size}")

            img_bytes = pil_to_bytes(page_img)
            prompt    = build_prompt(page_num, total_pages, uploaded.name)
            result    = call_gemini(client, img_bytes, prompt)

            ext = result.get("extracted", {})
            for k in all_ext:
                v = ext.get(k, [])
                if isinstance(v, list): all_ext[k].extend(v)

            iss_list = result.get("issues", [])
            if not isinstance(iss_list, list): iss_list = []
            for iss in iss_list:
                if isinstance(iss, dict) and iss.get("issue"):
                    all_issues.append({**iss, "file": uploaded.name, "page": page_num})

            highs = sum(1 for i in iss_list if i.get("severity")=="HIGH")
            if iss_list:
                log(f"&nbsp;&nbsp;&nbsp;⚠ {len(iss_list)} issue(s)"
                    + (f" — **{highs} HIGH**" if highs else "")
                    + f" `[{MODEL}]`")
            else:
                log(f"&nbsp;&nbsp;&nbsp;✅ No issues on this page `[{MODEL}]`")

            if idx < total_pages - 1:
                time.sleep(API_DELAY)

        progress_bar.progress(1.0)

        # 4. Deduplicate
        all_issues = dedupe_issues(all_issues)
        for k in all_ext: all_ext[k] = dedupe_list(all_ext[k])

        # 5. Build Excel
        xlsx_bytes = export_excel_bytes(all_issues, all_ext, meta)
        json_bytes = json.dumps(
            {"meta": meta, "issues": all_issues, "extracted": all_ext},
            indent=2
        ).encode()

        st.session_state.results = {
            "issues":     all_issues,
            "extracted":  all_ext,
            "meta":       meta,
            "xlsx":       xlsx_bytes,
            "pdf":        pdf_bytes,
            "json":       json_bytes,
            "filename":   Path(uploaded.name).stem,
        }
        status_box.success(f"✅ Analysis complete — {len(all_issues)} unique issues found.")
        log(f"<br>✅ **Done.** {len(all_issues)} unique issues after deduplication.")

    except Exception as e:
        status_box.error(f"❌ Error: {e}")
        log(f"❌ **Error:** {e}")

    finally:
        st.session_state.running = False

# ─── MAIN: RESULTS ────────────────────────────────────────────────────────────
R = st.session_state.results
if R:
    issues    = R["issues"]
    extracted = R["extracted"]
    meta      = R["meta"]

    st.markdown("---")
    st.markdown("## 📊 QC Results")

    # Summary cards
    sc = {s: sum(1 for i in issues if i.get("severity")==s) for s in SEV_ORDER}
    card_cols = st.columns(5)
    for col, (sev, count) in zip(card_cols, list(sc.items()) + [("TOTAL", len(issues))]):
        cls = f"sev-{sev}"
        col.markdown(
            f'<div class="sev-card {cls}">'
            f'<div class="sev-num">{count}</div>'
            f'<div class="sev-lbl">{sev}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Download buttons
    dl1, dl2, _ = st.columns([1, 1, 3])
    base = f"SSP_QC_{meta['drawing']}_{date.today().isoformat()}"
    dl1.download_button(
        "⬇ Download Marked-up PDF",
        data=export_pdf_markup_bytes(R.get("issues",[]), R.get("pdf", b""), dpi=dpi),
        file_name=f"{base}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )
    dl2.download_button(
        "⬇ Download JSON",
        data=R["json"],
        file_name=f"{base}.json",
        mime="application/json",
        use_container_width=True,
    )

    st.markdown("---")
    st.markdown("## 🗂 Issues")

    # Severity filter
    filter_sev = st.multiselect(
        "Filter by severity",
        options=SEV_ORDER,
        default=SEV_ORDER,
    )

    filtered = [i for i in issues if i.get("severity") in filter_sev]

    if filtered:
        df = pd.DataFrame([{
            "#":              idx+1,
            "Severity":       i.get("severity",""),
            "Category":       i.get("category",""),
            "Element":        i.get("element",""),
            "Issue":          i.get("issue",""),
            "Recommendation": i.get("recommendation",""),
            "Page":           i.get("page",""),
        } for idx, i in enumerate(filtered)])

        def colour_severity(val):
            c = {"HIGH":"#E84040","MEDIUM":"#E8B020","LOW":"#3A9FFF","INFO":"#2DCA72"}.get(val,"")
            return f"color:{c}; font-weight:bold" if c else ""

        styled = (
            df.style
            .applymap(colour_severity, subset=["Severity"])
            .set_properties(**{"white-space":"pre-wrap"})
        )
        st.dataframe(styled, use_container_width=True, height=520)
    else:
        st.info("No issues match the selected filters.")

    st.markdown("---")
    st.markdown("## 🔎 Extracted Drawing Data")

    ext_cols = st.columns(2)
    sections = [
        ("Equipment Tags",         "equipment_tags"),
        ("Line Numbers",           "line_numbers"),
        ("Instrument Tags",        "instrument_tags"),
        ("Valve Tags",             "valve_tags"),
        ("Pipe Specs",             "pipe_specs"),
        ('⚠ Placeholder X" Sizes', "placeholder_sizes"),
    ]

    for i, (title, key) in enumerate(sections):
        items = extracted.get(key, [])
        col   = ext_cols[i % 2]
        with col.expander(f"{title} ({len(items)})", expanded=(key=="placeholder_sizes")):
            if items:
                st.markdown(
                    "\n".join(f"- `{item}`" for item in items)
                )
            else:
                st.caption("None found")

    # Drawing metadata recap
    st.markdown("---")
    with st.expander("📋 Review Metadata"):
        mc1, mc2 = st.columns(2)
        for (k,v), col in zip(meta.items(), [mc1,mc2,mc1,mc2,mc1,mc2,mc1,mc2]):
            col.markdown(f"**{k.title()}:** {v}")
        st.markdown(f"**Model:** {MODEL}  \n**Date:** {date.today()}")
